from typing import Optional
import zipfile
from pathlib import Path

from smolagents import Tool
from smolagents.models import MessageRole, Model

from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()


class SheetExtractorTool(Tool):
    name = "extract_sheet_features"
    description = """
Extract specialized features from Excel/spreadsheet files beyond basic content.
This tool supports color extraction and media extraction for Excel files.
Supports .xlsx and .xls files. For basic content extraction, use other text processing tools."""

    inputs = {
        "file_path": {
            "description": "The path to the Excel file you want to process. Must be a '.xlsx' or '.xls' file.",
            "type": "string",
        },
        "feature_type": {
            "description": "Type of feature to extract: 'colors' (cell background colors), 'media' (embedded images/files), or 'formats' (list supported formats)",
            "type": "string",
            "nullable": True,
        },
        "sheet_name": {
            "description": "[Optional]: Specific sheet name to process. If not provided, uses the active/first sheet.",
            "type": "string",
            "nullable": True,
        },
        "question": {
            "description": "[Optional]: Your question about the extracted features. Provide context about what you're looking for.",
            "type": "string",
            "nullable": True,
        },
    }
    output_type = "string"

    def __init__(self, model: Model =None, workspace_path: Optional[str] = None):
        super().__init__()
        from pathlib import Path
        self.model = model
        self.workspace_path = Path(workspace_path) if workspace_path else Path.cwd()

        # Create output directories
        self._media_output_dir = self.workspace_path / "extracted_media"
        self._media_output_dir.mkdir(exist_ok=True)

        self.supported_extensions = {".xlsx", ".xls"}
        self.max_file_size = 100 * 1024 * 1024  # 100MB

    def _validate_file_type(self, file_path: str):
        """Validate if the file type is a supported Excel format"""
        from pathlib import Path
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if path.suffix.lower() not in self.supported_extensions:
            raise ValueError(
                f"Unsupported file type. This tool only supports {', '.join(self.supported_extensions)} files."
            )

        # Check file size
        if path.stat().st_size > self.max_file_size:
            raise ValueError(
                f"File too large. Maximum size is {self.max_file_size / (1024*1024):.1f}MB"
            )

    def _extract_cell_colors(
        self, file_path: str, sheet_name: Optional[str] = None
    ) -> dict:
        from pathlib import Path
        from openpyxl import load_workbook
        """Extract cell background colors from Excel file using openpyxl."""
        try:
            path = Path(file_path)

            # Only works with xlsx files using openpyxl
            if path.suffix.lower() != ".xlsx":
                return {"error": "Color extraction only supported for .xlsx files"}

            wb = load_workbook(file_path)

            # Select sheet
            if sheet_name and sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.active

            if sheet is None:
                return {"error": "No valid worksheet found"}

            cell_colors = {}

            # Iterate through cells
            for row in range(
                1, min((sheet.max_row or 0) + 1, 101)
            ):  # Limit to 100 rows
                for col in range(
                    1, min((sheet.max_column or 0) + 1, 21)
                ):  # Limit to 20 columns
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value
                    cell_coord = f"{row},{col}"

                    # Get cell background color if available
                    bg_color = None
                    if (
                        cell.fill
                        and hasattr(cell.fill, "start_color")
                        and hasattr(cell.fill.start_color, "rgb")
                    ):
                        rgb_obj = cell.fill.start_color.rgb
                        if rgb_obj:
                            if hasattr(rgb_obj, "rgb"):
                                bg_color = rgb_obj.rgb
                            else:
                                bg_color = str(rgb_obj)

                            # Convert AARRGGBB format to standard hex
                            if (
                                bg_color
                                and isinstance(bg_color, str)
                                and len(bg_color) == 8
                            ):
                                bg_color = bg_color[2:]  # Remove alpha channel

                    cell_colors[cell_coord] = {"value": cell_value, "color": bg_color}

            return cell_colors

        except Exception as e:
            return {"error": f"Failed to extract colors: {str(e)}"}

    def _extract_embedded_media(self, file_path: str) -> list:
        """Extract embedded media from XLSX files."""
        import zipfile # 内部导入
        from pathlib import Path
        from openpyxl import load_workbook
        saved_media = []

        try:
            path = Path(file_path)

            if path.suffix.lower() != ".xlsx":
                return [{"error": "Media extraction only supported for .xlsx files"}]

            # Load workbook to extract images
            workbook = load_workbook(file_path, data_only=False)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Try to extract images from the sheet (if available)
                try:
                    if hasattr(sheet, "_images"):
                        images = getattr(sheet, "_images", None)
                        if images:
                            for idx, image in enumerate(images):
                                try:
                                    image_data = image._data()
                                    image_filename = (
                                        f"{path.stem}_{sheet_name}_image_{idx}.png"
                                    )
                                    image_path = self._media_output_dir / image_filename

                                    with open(image_path, "wb") as f:
                                        f.write(image_data)

                                    saved_media.append(
                                        {
                                            "type": "image",
                                            "path": str(image_path),
                                            "filename": image_filename,
                                            "sheet": sheet_name,
                                        }
                                    )
                                except Exception as e:
                                    saved_media.append(
                                        {
                                            "error": f"Failed to extract image {idx} from sheet {sheet_name}: {str(e)}"
                                        }
                                    )
                except AttributeError:
                    # Sheet doesn't have _images attribute, skip image extraction for this sheet
                    pass

            # Also try to extract from ZIP structure for additional media
            with zipfile.ZipFile(file_path, "r") as zip_file:
                media_files = [
                    f for f in zip_file.namelist() if f.startswith("xl/media/")
                ]

                for media_file in media_files:
                    try:
                        media_data = zip_file.read(media_file)
                        media_filename = f"{path.stem}_{Path(media_file).name}"
                        media_path = self._media_output_dir / media_filename

                        with open(media_path, "wb") as f:
                            f.write(media_data)

                        # Determine media type based on extension
                        media_ext = Path(media_file).suffix.lower()
                        if media_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]:
                            media_type = "image"
                        elif media_ext in [".mp3", ".wav", ".m4a"]:
                            media_type = "audio"
                        elif media_ext in [".mp4", ".avi", ".mov"]:
                            media_type = "video"
                        else:
                            media_type = "other"

                        saved_media.append(
                            {
                                "type": media_type,
                                "path": str(media_path),
                                "filename": media_filename,
                                "original_path": media_file,
                            }
                        )
                    except Exception as e:
                        saved_media.append(
                            {"error": f"Failed to extract media {media_file}: {str(e)}"}
                        )

        except Exception as e:
            saved_media.append(
                {"error": f"Failed to extract media from XLSX: {str(e)}"}
            )

        return saved_media

    def _list_supported_formats(self) -> str:
        """List all supported Excel formats."""
        supported_formats = {
            "XLSX": "Excel 2007+ format files (.xlsx) - Full support including images and colors",
            "XLS": "Excel 97-2003 format files (.xls) - Limited support, colors and media not available",
        }

        format_list = "\n".join(
            [
                f"**{format_name}**: {description}"
                for format_name, description in supported_formats.items()
            ]
        )

        return f"Supported Excel formats:\n\n{format_list}"

    def forward(
        self,
        file_path: str,
        feature_type: Optional[str] = None,
        sheet_name: Optional[str] = None,
        question: Optional[str] = None,
    ) -> str:
        from smolagents.models import MessageRole # 内部导入
        try:
            if feature_type is None:
                feature_type = "formats"
            # Validate file type for most operations
            if feature_type != "formats":
                self._validate_file_type(file_path)

            # Handle different feature types
            if feature_type == "colors":
                result = self._extract_cell_colors(file_path, sheet_name)
                if "error" in result:
                    return f"Error extracting colors: {result['error']}"

                # Count cells with colors
                colored_count = 0
                for coord in result:
                    info = result[coord]
                    if info.get("color") and info["color"] != "FFFFFF":
                        colored_count += 1
                
                base_response = f"Successfully extracted colors from {len(result)} cells. Found {colored_count} cells with non-default background colors."

                if not question:
                    return f"{base_response}\n\nColor data available for analysis."

                # Use model to answer question about colors
                messages = [
                    {
                        "role": MessageRole.SYSTEM,
                        "content": [
                            {
                                "type": "text",
                                "text": f"Here is cell color data from an Excel file:\n{str(result)[:2000]}\n"
                                "Answer the following question about the color patterns in the spreadsheet.",
                            }
                        ],
                    },
                    {
                        "role": MessageRole.USER,
                        "content": [{"type": "text", "text": f"Question: {question}"}],
                    },
                ]

                content = self.model(messages).content
                return f"{base_response}\n\nAnalysis: {content}"

            elif feature_type == "media":
                result = self._extract_embedded_media(file_path)
                
                # 修复 3: 这里的 item 在列表推导式中定义，确保逻辑严密
                errors = [it["error"] for it in result if isinstance(it, dict) and "error" in it]
                if errors: return f"Errors: {'; '.join(errors)}"

                media_files = [it.get("filename") for it in result if isinstance(it, dict) and "filename" in it]
                base_response = f"Extracted {len(media_files)} media files."

                if not question: return f"{base_response} Files: {media_files}"

                messages = [
                    {"role": MessageRole.SYSTEM, "content": [{"type": "text", "text": f"Media Data: {str(result)[:1500]}"}]},
                    {"role": MessageRole.USER, "content": [{"type": "text", "text": question}]}
                ]
                return f"{base_response}\nAnalysis: {self.model(messages).content}"

            elif feature_type == "formats":
                return self._list_supported_formats()

            else:
                return f"Unknown feature type: {feature_type}. Supported types: colors, media, formats"

        except Exception as e:
            return f"Error processing Excel file: {str(e)}"
