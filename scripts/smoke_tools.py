#!/usr/bin/env python3
"""Initialize and minimally exercise MAS Arena tools.

The script is intentionally conservative: it uses local fixture files for file
tools and reports network/API-dependent tools as SKIP when their environment is
not configured.
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import inspect
import json
import os
import re
import sys
import tempfile
import wave
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


@dataclass
class SmokeResult:
    name: str
    status: str
    detail: str


def _shorten(value: Any, limit: int = 240) -> str:
    text = str(value).replace("\n", "\\n")
    text = re.sub(r"sk-[A-Za-z0-9_*.-]+", "sk-***", text)
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def _write_fixtures(tmp_dir: Path) -> dict[str, str]:
    fixtures: dict[str, str] = {}

    text_path = tmp_dir / "sample.txt"
    text_path.write_text("MAS Arena smoke test text fixture.\nAnswer: 42\n", encoding="utf-8")
    fixtures["text"] = str(text_path)

    csv_path = tmp_dir / "sample.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["name", "score"])
        writer.writerow(["alpha", 1])
        writer.writerow(["beta", 2])
    fixtures["csv"] = str(csv_path)

    json_path = tmp_dir / "sample.json"
    json_path.write_text(json.dumps({"name": "alpha", "score": 1}), encoding="utf-8")
    fixtures["json"] = str(json_path)

    zip_path = tmp_dir / "sample.zip"
    with zipfile.ZipFile(zip_path, "w") as zf:
        zf.write(text_path, arcname="sample.txt")
        zf.write(csv_path, arcname="sample.csv")
    fixtures["zip"] = str(zip_path)

    try:
        from PIL import Image

        image_path = tmp_dir / "sample.png"
        image = Image.new("RGB", (16, 16), color=(255, 255, 255))
        image.save(image_path)
        fixtures["image"] = str(image_path)
    except Exception:
        pass

    try:
        from openpyxl import Workbook

        xlsx_path = tmp_dir / "sample.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "name"
        ws["B1"] = "score"
        ws.append(["alpha", 1])
        ws.append(["beta", 2])
        wb.save(xlsx_path)
        fixtures["xlsx"] = str(xlsx_path)
    except Exception:
        pass

    wav_path = tmp_dir / "sample.wav"
    with wave.open(str(wav_path), "wb") as wav:
        wav.setnchannels(1)
        wav.setsampwidth(2)
        wav.setframerate(16000)
        wav.writeframes(b"\x00\x00" * 1600)
    fixtures["audio"] = str(wav_path)

    return fixtures


def _invoke(tool: Any, **kwargs: Any) -> Any:
    output = tool.forward(**kwargs)
    if inspect.isawaitable(output):
        return asyncio.run(output)
    return output


def _is_missing_config_output(text: str) -> bool:
    lowered = text.lower()
    if lowered.startswith("skip:"):
        return True
    return any(
        marker in lowered
        for marker in (
            "missing",
            "not available",
            "environment variable is required",
            "set jina_api_key",
            "api key",
        )
    )


def _is_tool_error_output(text: str) -> bool:
    lowered = text.lower()
    return lowered.startswith("error ") or lowered.startswith("error:")


def _run_case(
    name: str,
    tool_factory: Callable[[], Any],
    call: Callable[[Any], Any],
    allow_config_skip: bool = False,
) -> SmokeResult:
    try:
        tool = tool_factory()
    except Exception as exc:
        return SmokeResult(name, "FAIL", f"init failed: {_shorten(exc)}")

    try:
        output = call(tool)
    except Exception as exc:
        detail = _shorten(exc)
        if allow_config_skip and _is_missing_config_output(detail):
            return SmokeResult(name, "SKIP", detail)
        return SmokeResult(name, "FAIL", f"call failed: {detail}")

    detail = _shorten(output)
    if allow_config_skip and _is_missing_config_output(detail):
        return SmokeResult(name, "SKIP", detail)
    if _is_tool_error_output(detail):
        return SmokeResult(name, "FAIL", detail)
    return SmokeResult(name, "PASS", detail)


def run_smoke(include_network: bool = False, include_browser: bool = False) -> list[SmokeResult]:
    from mas_arena.tools import ALL_EXTERNAL_TOOLS, SimpleCrawler

    with tempfile.TemporaryDirectory(prefix="mas_arena_tool_smoke_") as tmp:
        tmp_dir = Path(tmp)
        fixtures = _write_fixtures(tmp_dir)
        extract_dir = tmp_dir / "unzipped"

        crawler = SimpleCrawler()
        cases: dict[str, tuple[Callable[[], Any], Callable[[Any], Any], bool]] = {
            "final_answer": (
                lambda: ALL_EXTERNAL_TOOLS["final_answer"](),
                lambda tool: _invoke(tool, answer="smoke-ok"),
                False,
            ),
            "python_interpreter": (
                lambda: ALL_EXTERNAL_TOOLS["python_interpreter"](),
                lambda tool: _invoke(tool, code="print(2 + 2)"),
                False,
            ),
            "text_extractor": (
                lambda: ALL_EXTERNAL_TOOLS["text_extractor"](),
                lambda tool: _invoke(tool, file_path=fixtures["text"], max_length=200),
                False,
            ),
            "csv_extractor": (
                lambda: ALL_EXTERNAL_TOOLS["csv_extractor"](),
                lambda tool: _invoke(
                    tool,
                    file_path=fixtures["csv"],
                    output_format="markdown",
                    max_rows=3,
                    include_statistics=True,
                ),
                False,
            ),
            "markdown_converter": (
                lambda: ALL_EXTERNAL_TOOLS["markdown_converter"](),
                lambda tool: _invoke(tool, file_path=fixtures["json"]),
                False,
            ),
            "zip_extractor": (
                lambda: ALL_EXTERNAL_TOOLS["zip_extractor"](),
                lambda tool: _invoke(
                    tool,
                    file_path=fixtures["zip"],
                    extract_dir=str(extract_dir),
                ),
                False,
            ),
            "sheet_extractor": (
                lambda: ALL_EXTERNAL_TOOLS["sheet_extractor"](),
                lambda tool: _invoke(
                    tool,
                    file_path=fixtures.get("xlsx", ""),
                    feature_type="formats",
                ),
                False,
            ),
            "visual_inspector": (
                lambda: ALL_EXTERNAL_TOOLS["visual_inspector"](),
                lambda tool: _invoke(tool, file_path=fixtures.get("image", ""), question=None),
                True,
            ),
            "audio_inspector": (
                lambda: ALL_EXTERNAL_TOOLS["audio_inspector"](),
                lambda tool: _invoke(tool, file_path=fixtures["audio"], question=None),
                True,
            ),
            "crawler_read": (
                lambda: ALL_EXTERNAL_TOOLS["crawler_read"](crawler=crawler, read_type="simple"),
                lambda tool: _invoke(tool, url="https://example.com"),
                True,
            ),
            "crawler_archive_search": (
                lambda: ALL_EXTERNAL_TOOLS["crawler_archive_search"](
                    crawler=crawler,
                    read_type="simple",
                ),
                lambda tool: _invoke(tool, url="https://example.com", date=None),
                True,
            ),
            "search": (
                lambda: ALL_EXTERNAL_TOOLS["search"](),
                lambda tool: _invoke(tool, query="MAS Arena smoke test"),
                True,
            ),
            "wikipedia_search": (
                lambda: ALL_EXTERNAL_TOOLS["wikipedia_search"](),
                lambda tool: _invoke(tool, query="Python (programming language)", max_sentences=1),
                True,
            ),
            "browser": (
                lambda: ALL_EXTERNAL_TOOLS["browser"](),
                lambda tool: _invoke(tool, action="get_url", url=None),
                True,
            ),
        }

        results: list[SmokeResult] = []
        for name in ALL_EXTERNAL_TOOLS:
            if name in {"search", "wikipedia_search"} and not include_network:
                results.append(SmokeResult(name, "SKIP", "network smoke disabled; pass --include-network"))
                continue
            if name == "browser" and not include_browser:
                results.append(SmokeResult(name, "SKIP", "browser smoke disabled; pass --include-browser"))
                continue

            case = cases.get(name)
            if case is None:
                results.append(SmokeResult(name, "FAIL", "no smoke case defined"))
                continue

            factory, call, allow_config_skip = case
            results.append(_run_case(name, factory, call, allow_config_skip))

        return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Smoke-test MAS Arena tools.")
    parser.add_argument(
        "--include-network",
        action="store_true",
        help="Call network-dependent tools such as search and Wikipedia.",
    )
    parser.add_argument(
        "--include-browser",
        action="store_true",
        help="Call the browser tool. This may require Playwright/browser dependencies.",
    )
    args = parser.parse_args()

    results = run_smoke(
        include_network=args.include_network,
        include_browser=args.include_browser,
    )

    width = max(len(result.name) for result in results)
    counts = {"PASS": 0, "SKIP": 0, "FAIL": 0}
    for result in results:
        counts[result.status] += 1
        print(f"{result.status:<4} {result.name:<{width}} {result.detail}")

    print(
        f"\nSummary: {counts['PASS']} passed, {counts['SKIP']} skipped, "
        f"{counts['FAIL']} failed."
    )
    return 1 if counts["FAIL"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
